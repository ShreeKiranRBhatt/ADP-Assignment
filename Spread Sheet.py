# setting the font style of the cells
import openpyxl 
  # import Font function from openpyxl 
from openpyxl.styles import Font   
wb = openpyxl.Workbook() 
sheet = wb.active   
sheet.cell(row = 1, column = 1).value = "Shreekiran"
  # set the size of the cell to 24 
sheet.cell(row = 1, column = 1).font = Font(size = 24 ) 
 sheet.cell(row = 2, column = 2).value = "Raj"
 # set the font style to italic 
sheet.cell(row = 2, column = 2).font = Font(size = 24, italic = True) 
  sheet.cell(row = 3, column = 3).value = "Ram"
  # set the font style to bold 
sheet.cell(row = 3, column = 3).font = Font(size = 24, bold = True) 
  sheet.cell(row = 4, column = 4).value = "Krishna"
  # set the font name to 'Times New Roman' 
sheet.cell(row = 4, column = 4).font = Font(size = 24, name = 'Times New Roman') 
  wb.save('styles.xlsx')


# Adjusting rows and columns
# import openpyxl module 
import openpyxl 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute.  
sheet = wb.active 
  
# writing to the specified cell 
sheet.cell(row = 1, column = 1).value = ' hello '
  
sheet.cell(row = 2, column = 2).value = ' everyone '
  
# set the height of the row 
sheet.row_dimensions[1].height = 70
  
# set the width of the column 
sheet.column_dimensions['B'].width = 20
  
# save the file 
wb.save('dimension.xlsx')


# Setting Charts 
import openpyxl module 
import openpyxl 
  
# import BarChart class from openpyxl.chart sub_module 
from openpyxl.chart import BarChart,Reference 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet  
# from the active attribute. 
sheet = wb.active 
  
# write o to 9 in 1st column of the active sheet 
for i in range(10): 
    sheet.append([i]) 
  
# create data for plotting 
values = Reference(sheet, min_col = 1, min_row = 1, 
                         max_col = 1, max_row = 10) 
  
# Create object of BarChart class 
chart = BarChart() 
  
# adding data to the Bar chart object 
chart.add_data(values) 
  
# set the title of the chart 
chart.title = " BAR-CHART "
  
# set the title of the x-axis 
chart.x_axis.title = " X_AXIS "
  
# set the title of the y-axis 
chart.y_axis.title = " Y_AXIS "
  
# add chart to the sheet 
# the top-left corner of a chart 
# is anchored to cell E2 . 
sheet.add_chart(chart, "E2") 
  
# save the file 
wb.save("barChart.xlsx")



# updating a spreadsheet
# import openpyxl module 
import openpyxl 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute.  
sheet = wb.active 
  
# Once have the Worksheet object, 
# one can get its name from the 
# title attribute. 
sheet_title = sheet.title 
  
print("active sheet title: " + sheet_title)

# import openpyxl module 
import openpyxl 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute 
sheet = wb.active 
  
# One can change the name of the title 
sheet.title = "sheet1"
  
print("sheet name is renamed as: " + sheet.title)



# setting formulas
# import openpyxl module 
import openpyxl 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute. 
sheet = wb.active 
  
# writing to the cell of an excel sheet 
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = 400
sheet['A4'] = 500
sheet['A5'] = 600
  
# The value in cell A7 is set to a formula  
# that sums the values in A1, A2, A3, A4, A5 . 
sheet['A7'] = '= SUM(A1:A5)'
  
# save the file 
wb.save("sum.xlsx")



